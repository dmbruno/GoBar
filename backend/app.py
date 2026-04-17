import os
import smtplib
import time
import base64
import json
import threading
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage

from flask import Flask, request, jsonify, Response
from flask_cors import CORS
from dotenv import load_dotenv
from openpyxl import load_workbook
from io import BytesIO

load_dotenv()

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

# ──────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────
GMAIL_USER = os.getenv("GMAIL_USER")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")
EMAIL_DELAY = int(os.getenv("EMAIL_DELAY", 2))
MAX_EMAILS_PER_BATCH = int(os.getenv("MAX_EMAILS_PER_BATCH", 450))

# In-memory job tracking (single user app, no DB needed)
jobs = {}


# ──────────────────────────────────────────────
# Email template
# ──────────────────────────────────────────────
def build_html_email(subject, body_text, has_image=False):
    """Builds a professional HTML email template."""
    image_block = ""
    if has_image:
        image_block = """
        <tr>
          <td style="padding: 0 0 20px 0; text-align: center;">
            <img src="cid:campaign_image"
                 style="max-width: 100%; height: auto; border-radius: 8px;"
                 alt="Imagen de campaña" />
          </td>
        </tr>
        """

    # Convert newlines to <br> for the body
    body_html = body_text.replace("\n", "<br>")

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
      <meta charset="utf-8">
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="margin: 0; padding: 0; background-color: #f4f4f4; font-family: Arial, Helvetica, sans-serif;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background-color: #f4f4f4;">
        <tr>
          <td align="center" style="padding: 20px 0;">
            <table role="presentation" width="600" cellpadding="0" cellspacing="0"
                   style="background-color: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">

              <!-- Header -->
              <tr>
                <td style="background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); padding: 30px 40px; text-align: center;">
                  <h1 style="color: #ffffff; margin: 0; font-size: 24px; font-weight: bold;">
                    {subject}
                  </h1>
                </td>
              </tr>

              <!-- Image -->
              {image_block}

              <!-- Body -->
              <tr>
                <td style="padding: 30px 40px;">
                  <p style="color: #333333; font-size: 16px; line-height: 1.6; margin: 0;">
                    {body_html}
                  </p>
                </td>
              </tr>

              <!-- Footer -->
              <tr>
                <td style="background-color: #f8f9fa; padding: 20px 40px; text-align: center; border-top: 1px solid #e9ecef;">
                  <p style="color: #999999; font-size: 12px; margin: 0;">
                    Este correo fue enviado por GoBar. Si no deseas recibir más correos,
                    por favor respondé a este email con el asunto "DESUSCRIBIR".
                  </p>
                </td>
              </tr>

            </table>
          </td>
        </tr>
      </table>
    </body>
    </html>
    """
    return html


def send_single_email(to_email, to_name, subject, html_content, image_data=None):
    """Send a single email via Gmail SMTP."""
    msg = MIMEMultipart("related")
    msg["From"] = GMAIL_USER
    msg["To"] = to_email
    msg["Subject"] = subject

    # Personalizar saludo si hay nombre
    if to_name:
        personalized_html = html_content.replace("{{nombre}}", to_name)
    else:
        personalized_html = html_content.replace("{{nombre}}", "")

    msg.attach(MIMEText(personalized_html, "html"))

    # Attach image inline if provided
    if image_data:
        img = MIMEImage(image_data["bytes"], _subtype=image_data["subtype"])
        img.add_header("Content-ID", "<campaign_image>")
        img.add_header("Content-Disposition", "inline", filename=image_data["filename"])
        msg.attach(img)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.send_message(msg)


def process_email_job(job_id, contacts, subject, body_text, image_data):
    """Background thread that sends all emails for a job."""
    job = jobs[job_id]
    job["status"] = "sending"

    html_template = build_html_email(subject, body_text, has_image=image_data is not None)

    for i, contact in enumerate(contacts):
        if job.get("cancelled"):
            job["status"] = "cancelled"
            return

        email = contact["email"]
        name = contact.get("name", "")

        try:
            send_single_email(email, name, subject, html_template, image_data)
            job["sent"] += 1
            job["log"].append({"email": email, "status": "ok", "name": name})
        except Exception as e:
            job["failed"] += 1
            job["log"].append({"email": email, "status": "error", "error": str(e), "name": name})

        job["progress"] = int(((i + 1) / job["total"]) * 100)

        # Delay between emails to avoid spam filters
        if i < len(contacts) - 1:
            time.sleep(EMAIL_DELAY)

    job["status"] = "completed"


# ──────────────────────────────────────────────
# Routes
# ──────────────────────────────────────────────

@app.route("/api/health", methods=["GET"])
def health():
    """Health check."""
    configured = bool(GMAIL_USER and GMAIL_APP_PASSWORD)
    return jsonify({"status": "ok", "gmail_configured": configured})


@app.route("/api/parse-excel", methods=["POST"])
def parse_excel():
    """Parse uploaded Excel/CSV and return list of contacts with emails."""
    if "file" not in request.files:
        return jsonify({"error": "No se envió ningún archivo"}), 400

    file = request.files["file"]
    filename = file.filename.lower()

    try:
        contacts = []

        if filename.endswith((".xlsx", ".xls")):
            wb = load_workbook(BytesIO(file.read()), read_only=True)
            ws = wb.active

            # Find email column (search in first row)
            headers = [str(cell.value or "").lower().strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            email_col = None
            name_col = None
            lastname_col = None

            for idx, header in enumerate(headers):
                if "email" in header or "correo" in header or "e-mail" in header or "mail" in header:
                    email_col = idx
                if header in ("nombre", "name", "first_name", "first name"):
                    name_col = idx
                if header in ("apellido", "lastname", "last_name", "last name", "surname"):
                    lastname_col = idx

            if email_col is None:
                return jsonify({"error": "No se encontró una columna de email en el archivo. Asegurate de tener una columna con 'Email' en el encabezado."}), 400

            for row in ws.iter_rows(min_row=2, values_only=True):
                email = str(row[email_col] or "").strip()
                if not email or "@" not in email:
                    continue

                name_parts = []
                if name_col is not None and row[name_col]:
                    name_parts.append(str(row[name_col]).strip())
                if lastname_col is not None and row[lastname_col]:
                    name_parts.append(str(row[lastname_col]).strip())

                contacts.append({
                    "email": email,
                    "name": " ".join(name_parts) if name_parts else ""
                })

            wb.close()

        elif filename.endswith(".csv"):
            import csv
            import io
            content = file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))

            email_field = None
            name_field = None
            lastname_field = None

            for field in reader.fieldnames or []:
                fl = field.lower().strip()
                if "email" in fl or "correo" in fl or "mail" in fl:
                    email_field = field
                if fl in ("nombre", "name", "first_name", "first name"):
                    name_field = field
                if fl in ("apellido", "lastname", "last_name", "last name"):
                    lastname_field = field

            if not email_field:
                return jsonify({"error": "No se encontró columna de email en el CSV"}), 400

            for row in reader:
                email = (row.get(email_field) or "").strip()
                if not email or "@" not in email:
                    continue

                name_parts = []
                if name_field and row.get(name_field):
                    name_parts.append(row[name_field].strip())
                if lastname_field and row.get(lastname_field):
                    name_parts.append(row[lastname_field].strip())

                contacts.append({
                    "email": email,
                    "name": " ".join(name_parts) if name_parts else ""
                })
        else:
            return jsonify({"error": "Formato no soportado. Usá .xlsx, .xls o .csv"}), 400

        # Remove duplicates by email
        seen = set()
        unique_contacts = []
        for c in contacts:
            if c["email"].lower() not in seen:
                seen.add(c["email"].lower())
                unique_contacts.append(c)

        return jsonify({
            "contacts": unique_contacts,
            "total": len(unique_contacts),
            "duplicates_removed": len(contacts) - len(unique_contacts)
        })

    except Exception as e:
        return jsonify({"error": f"Error al procesar el archivo: {str(e)}"}), 500


@app.route("/api/send", methods=["POST"])
def send_emails():
    """Start sending emails in background."""
    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        return jsonify({"error": "Gmail no está configurado. Revisá las variables de entorno GMAIL_USER y GMAIL_APP_PASSWORD."}), 500

    # Parse form data
    subject = request.form.get("subject", "").strip()
    body = request.form.get("body", "").strip()
    contacts_json = request.form.get("contacts", "[]")

    if not subject:
        return jsonify({"error": "El asunto es obligatorio"}), 400
    if not body:
        return jsonify({"error": "El cuerpo del correo es obligatorio"}), 400

    try:
        contacts = json.loads(contacts_json)
    except json.JSONDecodeError:
        return jsonify({"error": "Lista de contactos inválida"}), 400

    if not contacts:
        return jsonify({"error": "No hay contactos para enviar"}), 400

    # Limit batch size
    if len(contacts) > MAX_EMAILS_PER_BATCH:
        contacts = contacts[:MAX_EMAILS_PER_BATCH]
        batch_limited = True
    else:
        batch_limited = False

    # Process image if uploaded
    image_data = None
    if "image" in request.files:
        img_file = request.files["image"]
        if img_file.filename:
            img_bytes = img_file.read()
            ext = img_file.filename.rsplit(".", 1)[-1].lower()
            subtype_map = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "gif": "gif", "webp": "webp"}
            image_data = {
                "bytes": img_bytes,
                "subtype": subtype_map.get(ext, "png"),
                "filename": img_file.filename
            }

    # Create job
    job_id = f"job_{int(time.time() * 1000)}"
    jobs[job_id] = {
        "status": "queued",
        "total": len(contacts),
        "sent": 0,
        "failed": 0,
        "progress": 0,
        "log": [],
        "batch_limited": batch_limited,
        "original_total": len(json.loads(contacts_json))
    }

    # Start background thread
    thread = threading.Thread(
        target=process_email_job,
        args=(job_id, contacts, subject, body, image_data),
        daemon=True
    )
    thread.start()

    return jsonify({"job_id": job_id, "total": len(contacts), "batch_limited": batch_limited})


@app.route("/api/job/<job_id>", methods=["GET"])
def get_job_status(job_id):
    """Get status of an email sending job."""
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job no encontrado"}), 404

    return jsonify({
        "status": job["status"],
        "total": job["total"],
        "sent": job["sent"],
        "failed": job["failed"],
        "progress": job["progress"],
        "log": job["log"][-20:],  # Last 20 entries
        "batch_limited": job.get("batch_limited", False)
    })


@app.route("/api/job/<job_id>/cancel", methods=["POST"])
def cancel_job(job_id):
    """Cancel a running job."""
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job no encontrado"}), 404

    job["cancelled"] = True
    return jsonify({"message": "Cancelación solicitada"})


@app.route("/api/test-connection", methods=["POST"])
def test_connection():
    """Test Gmail SMTP connection."""
    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        return jsonify({"error": "Gmail no está configurado"}), 500

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=10) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        return jsonify({"status": "ok", "message": f"Conexión exitosa con {GMAIL_USER}"})
    except smtplib.SMTPAuthenticationError:
        return jsonify({"error": "Credenciales inválidas. Verificá tu App Password."}), 401
    except Exception as e:
        return jsonify({"error": f"Error de conexión: {str(e)}"}), 500


# ──────────────────────────────────────────────
# Run
# ──────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=True, port=5001)
